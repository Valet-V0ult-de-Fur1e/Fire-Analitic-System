import calendar
import openpyxl
import os
# years_dict = {
#     "Нижне-Енисейское.xlsx": "Красноярский край - Енисейский район .xlsx",
#     "Тунгусско-Чунское.xlsx": "Красноярский край - Тунгусско-Чунский район .xlsx",
#     "Чунское.xlsx": "Иркутская область - Чунский район .xlsx",
#     "Кодинское.xlsx": "Ханты-Мансийский автономный округ - Югра - Кондинский район .xlsx"
# }
years_dict = {
    "fires.xlsx": "Хабаровский край - Верхнебуреинский район .xlsx"
}


def get_transform_data_from_num(num, year1):
    dop_months_sleap_year = [31, 60, 91, 121, 152, 182, 213, 244, 274, 305, 335, 366]
    dop_months_nosleap_year = [31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334, 365]
    year_list = dop_months_sleap_year if calendar.isleap(year1) else dop_months_nosleap_year
    for range_data_index in range(12):
        if num <= year_list[range_data_index]:
            format_day = num - year_list[range_data_index - 1] if range_data_index != 0 else num
            month_new = "0" + str(range_data_index + 1) if len(str(range_data_index + 1)) == 1 else str(range_data_index + 1)
            day_new = "0" + str(format_day) if len(str(format_day)) == 1 else str(format_day)
            return f"{year1}.{month_new}.{day_new}"


main_dict = {
    # "Нижне-Енисейское.xlsx": dict(),
    # "Тунгусско-Чунское.xlsx": dict(),
    # "Чунское.xlsx": dict(),
    # "Кодинское.xlsx": dict()
    "fires.xlsx": dict()
}
files_name = {
    "fires.xlsx": "Верхнебуриинский.xlsx"
    # "Нижне-Енисейское.xlsx": "Енисейский район.xlsx",
    # "Тунгусско-Чунское.xlsx": "Тунгусско-Чунский район.xlsx",
    # "Чунское.xlsx": "Чунский район.xlsx",
    # "Кодинское.xlsx": "Югра-Кондинский район.xlsx"
}
for country in years_dict:
    main_dict[country] = dict()
    work_book = openpyxl.load_workbook(country)
    work_list = work_book[work_book.sheetnames[-1]]
    for i in work_list:
        main_dict[country][i[0].value] = [i[1].value]
    work_book_meteo = openpyxl.load_workbook(years_dict[country])
    work_book = None
    work_list = work_book_meteo[work_book_meteo.sheetnames[1]]
    for parametr_name in work_list[1][1:]:
        main_dict[country]["день"].append(parametr_name.value)
    for year in work_book_meteo.sheetnames[1:]:
        work_list = work_book_meteo[year]
        time_range = work_list[2:366] if calendar.isleap(int(year)) else work_list[2:365]
        for day in time_range:
            day_data_true = get_transform_data_from_num(int(day[0].value), int(year))
            for item in day[1:]:
                if day_data_true in main_dict[country].keys():
                    main_dict[country][day_data_true].append(int(item.value))
os.mkdir('rdn')
for country in main_dict:
    output_workbook = openpyxl.Workbook()
    worked_list = output_workbook.create_sheet('all years')
    for day_data in main_dict[country]:
        parameters_list = [day_data]
        for parameters in main_dict[country][day_data]:
            parameters_list.append(parameters)
        worked_list.append(parameters_list[:19])
    output_workbook.save(filename=f"rdn/{country}")
